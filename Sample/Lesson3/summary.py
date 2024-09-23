import openpyxl
import pandas as pd
import datetime
# 追加
import sys

args = sys.argv

year = int(args[1][:4])
month = int(args[1][4:])

def to_minute(val):
    return val.hour*60 + val.minute

wb_members = openpyxl.load_workbook(
    '従業員集約.xlsx', data_only=True)
names = wb_members.sheetnames

data = []
for name in names:
    df = pd.read_excel('従業員集約.xlsx', sheet_name=name)
    start_date = datetime.date(year, month, 1)
    end_date = (datetime.date(year, month+1, 1)
                - datetime.timedelta(days=1))

    df_filtered = df[(df['日付'].dt.date >= start_date) 
                     & (df['日付'].dt.date <= end_date)]

    df_filtered_fillna = df_filtered['作業時間'].fillna(
        datetime.time())
    working_hours = df_filtered_fillna.apply(to_minute)
    total_working_hour = working_hours.sum()/60

    datum = {
        '氏名': name,
        '合計作業時間': total_working_hour
    }
    data.append(datum)

df_summary = pd.DataFrame(data)

wb_summary = openpyxl.load_workbook('集計.xlsx', data_only=True)
ws_summary = wb_summary.create_sheet(title=f'{year}年{month}月分')

ws_summary['A1'] = '氏名'
ws_summary['B1'] = '合計作業時間'

ws_summary.column_dimensions['A'].width = 15
ws_summary.column_dimensions['B'].width = 15

for index, row in df_summary.iterrows():
    ws_summary.cell(row=index+2, column=1).value = row['氏名']
    ws_summary.cell(row=index+2, column=2).value = row['合計作業時間']

wb_summary.save('集計.xlsx')
print(f'{year}年{month}月分の集計完了')