import openpyxl
from glob import glob
import datetime

# 追加
import sys
args = sys.argv

# 　＜---必要な情報の取得---＞
file_paths = glob(f'{args[1]}/*')
# 変更部分：従業員集約.xlsxの読み込みを上部に移動
wb_members = openpyxl.load_workbook('従業員集約.xlsx', data_only=True)
for file_path in file_paths:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb['timesheet']

    name = ws['C3'].value
    datum = []
    for row in ws['B6:F12']:
        row_datum = []
        for cell in row:
            basis_date = datetime.datetime(1899, 12, 30, 0, 0)
            if cell.value == basis_date:
                row_datum.append(None)            
            else:
                row_datum.append(cell.value)
        datum.append(row_datum)

    # 　＜---勤怠情報の転記---＞    
    # シートが作成ずみかどうかで条件分岐
    if name not in wb_members.sheetnames:
        # シートが作成されていなかったら作成し、カラム名、カラム幅を設定
        wb_members.create_sheet(title=name)
        ws_members = wb_members[name]

        ws_members['A1'].value = '日付'
        ws_members['B1'].value = '出勤時間'
        ws_members['C1'].value = '退勤時間'
        ws_members['D1'].value = '休憩時間'
        ws_members['E1'].value = '作業時間'

        ws_members.column_dimensions['A'].width = 15
        ws_members.column_dimensions['B'].width = 15
        ws_members.column_dimensions['C'].width = 15    
        ws_members.column_dimensions['D'].width = 15    
        ws_members.column_dimensions['E'].width = 15    

    else:
        # シートが作成されていたら、そのシートを読み込む
        ws_members = wb_members[name]

    max_row = ws_members.max_row

    for row_num in range(1, 8):
        for column_num in range(1,6):
            if column_num == 1:
                ws_members.cell(
                    row=max_row+row_num, column=column_num
                ).number_format = 'yyyy/m/d'
            else:
                ws_members.cell(
                    row=max_row+row_num, column=column_num
                ).number_format = 'h:mm'
            ws_members.cell(
                row=max_row+row_num, column=column_num
            ).value = datum[row_num-1][column_num-1]

    wb_members.save('従業員集約.xlsx')

# 追加
print('集約完了')